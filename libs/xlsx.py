import openpyxl
from openpyxl.styles import Font


class SpreadsheetManager ():
    """ Manage local spread sheets
    """

    def __init__(self, file_name):

        self.file_name = file_name
        try:
            self.wb = openpyxl.load_workbook(self.file_name)
        except Exception:
            self.wb = openpyxl.Workbook()
            self.wb.save(filename=self.file_name)
        self.current_sheet = None

    def get_sheets(self) -> list:
        """ Return all sheets in current workbook

        Returns:
            list: List of all sheets in current workbook
        """

        return self.wb.sheetnames

    def clean_workbook(self):
        """ Delete all sheets in current workbook
        """

        for sheet in self.wb.sheetnames:
            self.delete_sheet(sheet)

    def delete_sheet(self, sheet_name: str):
        """ Delete a specific sheet in current workbook

        Args:
            sheet_name (str): Name of the sheet to be deleted
        """
        
        sheet_obj = self.wb[sheet_name]
        self.wb.remove(sheet_obj)

    def create_set_sheet(self, sheet_name: str):
        """ Create a new sheet in current workbook (if not exists)
        and set it as current sheet

        Args:
            sheet_name (str): Name of the new sheet
        """
        
        if sheet_name in self.get_sheets():
            self.set_sheet(sheet_name)
        else:
            self.wb.create_sheet(sheet_name)
            self.set_sheet(sheet_name)

    def set_sheet(self, sheet_name: str):
        """ Set a specific sheet as current sheet

        Args:
            sheet_name (str): Name of the sheet to be set as current
        """

        self.current_sheet = self.wb[sheet_name]

    def save(self):
        """ Save current workbook
        """

        self.wb.save(self.file_name)

    def write_cell(self, value: str = "", row: int = 1, column: int = 1):
        """ Write a value in a specific cell

        Args:
            value (str, optional): Value to be written. Defaults to "".
            row (int, optional): Row number. Defaults to 1.
            column (int, optional): Column number. Defaults to 1.
        """

        self.current_sheet.cell(row, column).value = value

    def write_data(self, data: list = [], start_row: int = 1, start_column: int = 1):
        """ Write a matrix of data in the current sheet

        Args:
            data (list, optional): Matrix of data. Defaults to [].
            start_row (int, optional): Row number to start writing. Defaults to 1.
            start_column (int, optional): Column number to start writing. Defaults to 1.
        """

        current_row = start_row
        current_column = start_column

        for row in data:

            for cell_value in row:

                cell_obj = self.current_sheet.cell(current_row, current_column)
                cell_obj.value = cell_value

                current_column += 1

            current_column = start_column
            current_row += 1

    def auto_width(self):
        """ Set corect width to each coumn in the current sheet
        """

        for col in self.current_sheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            self.current_sheet.column_dimensions[column].width = adjusted_width

    def format_range(self, start_cell: tuple = (1, 1), end_cell: tuple = (1, 1),
                     italic: bool = False, bold: bool = False, font_size: int = 8):
        """ Apply a specific style to a range of cells

        Args:
            start_cell (tuple, optional): Cell to start formatting. Defaults to (1, 1).
            end_cell (tuple, optional): Cell to end formatting. Defaults to (1, 1).
            italic (bool, optional): True if italic. Defaults to False.
            bold (bool, optional): True if bold. Defaults to False.
            font_size (int, optional): Font size. Defaults to 8.
        """

        # Create font style
        formated_font = Font(size=font_size, italic=italic, bold=bold)

        # Apply style
        current_row = start_cell[0]
        current_column = start_cell[1]

        for _ in range(start_cell[0], end_cell[0] + 1):

            for _ in range(start_cell[1], end_cell[1] + 1):

                cell_obj = self.current_sheet.cell(current_row, current_column)
                cell_obj.font = formated_font

                current_column += 1

            current_column = 1
            current_row += 1

    def get_data(self):
        """ Get all data from the current page """

        rows = self.current_sheet.max_row
        columns = self.current_sheet.max_column

        data = []
        for row in range(1, rows + 1):

            row_data = []
            for column in range(1, columns + 1):
                cell_data = self.current_sheet.cell(row, column).value
                row_data.append(cell_data)

            data.append(row_data)

        return data
